from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.chart.label import DataLabelList

wb = load_workbook("Stock_data_report.xlsx")
sheet = wb.active

min_col = sheet.min_column
max_col = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

chart = BarChart()

chart.title = "Stock Prices Comparison"
chart.x_axis.title = "Stock"
chart.y_axis.title = "Stock Prices"
chart.type = "col"
chart.width = 15
chart.height = 8
chart.style = 2

data = Reference(sheet, min_col=min_col +1, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col,min_row= min_row +1, max_row=max_row)

chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True
chart.dataLabels.position = "inEnd"
chart.y_axis.majorGridlines = None

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

sheet.add_chart(chart, "A15")

wb.save("Stock_data_report.xlsx")

