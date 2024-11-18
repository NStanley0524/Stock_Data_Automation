from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.chart.label import DataLabelList


wb = load_workbook("Stock_data_report.xlsx")
sheet = wb.active

min_col = sheet.min_column
max_col = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

bar = BarChart()

bar.title = "P.E Ratio Comparison"
bar.x_axis.title = "Stock"
bar.y_axis.title = "P.E.Ratio"
bar.type = "bar"
bar.width = 15
bar.height = 8
bar.style = 3

data = Reference(sheet, min_col=max_col, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col= min_col, min_row= min_row +1, max_row= max_row)

bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True
bar.dataLabels.position = "inEnd"
bar.y_axis.majorGridlines = None

bar.add_data(data, titles_from_data=True)
bar.set_categories(categories)
sheet.add_chart(bar, "K15")

wb.save("Stock_data_report.xlsx")

